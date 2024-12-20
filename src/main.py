import re
from openpyxl import load_workbook
import undetected_chromedriver as uc
import time
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from dotenv import load_dotenv
import os
from selenium.webdriver.common.keys import Keys

entradaExcel = r"C:\Users\diogo.lana\Desktop\ProjetoSenatran\data\BASE WAY TO GO SENATRAN.xlsx"

#Função rentorna infrações
def voltarInfra():

    campo_infracao_do_veiculo = WebDriverWait(navegador,10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > div.session-info-container > br-breadcrumbs > div > ul > li:nth-child(5) > a"))
    )
    navegador.execute_script("arguments[0].click();", campo_infracao_do_veiculo)

    try:
        modal_infracoes_aVencer = WebDriverWait(navegador, 15).until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-infracao > app-infracoes-list > app-infracoes-veiculo-list > app-infrator-list > div > div > app-infracao-lista > form > div.lista-infracoes.ng-star-inserted"))
        )
    except (NoSuchElementException, TimeoutException):
        navegador.refresh()
        time.sleep(10)

#Função volta pesquisa
def voltarPesquisa():

    campoConsultaInfra = WebDriverWait(navegador, 15).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > div > br-breadcrumbs > div > ul > li:nth-child(3) > a"))
    )

    navegador.execute_script("arguments[0].click();", campoConsultaInfra)

    botaoPorVeiculo = WebDriverWait(navegador, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "app-infracoes-list > div > div > div > div"))
    )

    navegador.execute_script("arguments[0].click();", botaoPorVeiculo)

    campoPlaca = WebDriverWait(navegador, 15).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "br-main-layout app-infracao > app-infracoes-list > app-infracoes-veiculo-list  form div.col-lg-3.col-md-4.col-sm-6 input"))
    )


# Configurar o Chrome com um User-Agent falso usando undetected-chromedriver
chrome_options = Options()
chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36")

# Inicializa o navegador com as opções e com o stealth mode ativado
navegador = uc.Chrome(options=chrome_options)
navegador.maximize_window()

time.sleep(5)

# Acessa o site especificado
navegador.get("https://portalservicos.senatran.serpro.gov.br/#/login")

#Aguarda até o elemento ficar visível (com um timeout de 60 segundos)
try:

    campoCPF = WebDriverWait(navegador, 60).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#accountId"))
    )
    campoCPF.send_keys(os.getenv("CPF"))

    botaoContinuarLogin = navegador.find_element(By.CSS_SELECTOR, "#enter-account-id")
    botaoContinuarLogin.click()

    campoSenha =  WebDriverWait(navegador, 1000).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#password"))
    )
    campoSenha.send_keys(os.getenv("SENHA"))

    botaoEntrar = navegador.find_element(By.CSS_SELECTOR, "#submit-button")
    botaoEntrar.click()

    telaInserirCNPJ = WebDriverWait(navegador,15).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-header"))
    )

    # Clica em "Representante Jurídico"
    representanteJuridico = navegador.find_element(By.CSS_SELECTOR, ".modal-content ul > li:nth-child(2) a")
    navegador.execute_script("arguments[0].click();", representanteJuridico)
    telaCnpj = WebDriverWait(navegador, 30).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div:nth-child(4) > div:nth-child(1) > div > h5"))
    )

    # Seleciona o CNPJ no dropdown
    select_cnpj = navegador.find_element(By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div:nth-child(4) > div:nth-child(2) > div > div > select")
    select = Select(select_cnpj)
    select.select_by_value(os.getenv("CnpjCliente"))

    time.sleep(1.5)

    # Clicar após selecionar o cliente
    botaoSelecionarCleinte = navegador.find_element(By.CSS_SELECTOR, "body > modal-container > div > div > div.modal-body.view-dashboard > div.modal-footer > div > button.br-button.primary.small.footer-button")
    botaoSelecionarCleinte.click()

    time.sleep(1)

    cokiies = navegador.find_element(By.CSS_SELECTOR, "#cookiebar > div.br-cookiebar.default > div > div > div > div.br-modal-footer.actions.justify-content-end > button.br-button.primary.small")
    navegador.execute_script("arguments[0].click();", cokiies)

    botaoInfracoes = WebDriverWait(navegador, 20).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-usuario > app-home > div > div.view-dashboard > ul > li:nth-child(2) > a"))
    ).click()
    
    time.sleep(1.5)

    botaoPorVeiculo = WebDriverWait(navegador, 20).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "app-infracoes-list > div > div > div > div"))
    ).click()


    #INICIA O EXCEL
    planilha = load_workbook(entradaExcel)

    #Define as instancias da planilha do Excel
    guia_veiculos = planilha['VEICULOS']
    guia_resultado_autos = planilha['RESULTADO AUTOS']

    #Carrega os cabeçalhos da planilha
    guia_veiculos['A1'] = "PLACA"
    guia_veiculos['B1'] = "RENAVAM"
    guia_veiculos['C1'] = "STATUS"
    guia_veiculos['D1'] = "QUANTIDADE DE MULTAS"

    #placa = []
    #renavam = []
    #status = []

    index = 0
    linhas = list(guia_veiculos.iter_rows(min_row=2, max_row=guia_veiculos.max_row))

    while index < len(linhas):
        row = linhas[index]

        placa_atual = row[0].value
        renavam_atual = row[1].value
        status_atual = row[2].value

        if status_atual is None:

            try:

                time.sleep(3)

                campoPlaca = WebDriverWait(navegador, 15).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "br-main-layout app-infracao > app-infracoes-list > app-infracoes-veiculo-list form div.col-lg-3.col-md-4.col-sm-6 input"))
                )
                campoPlaca.clear()
                campoPlaca.send_keys(placa_atual)

                # Aguarda que o valor seja inserido corretamente
                WebDriverWait(navegador, 10).until(
                    EC.text_to_be_present_in_element_value((By.CSS_SELECTOR, "br-main-layout app-infracao > app-infracoes-list > app-infracoes-veiculo-list form div.col-lg-3.col-md-4.col-sm-6 input"), placa_atual)
                )

                time.sleep(1.5)

                botaoPesquisarVeiculo = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-infracao > app-infracoes-list > app-infracoes-veiculo-list > div > div > app-infracao-veiculo-lista > form > div:nth-child(2) > div.col-lg-6.col-md-6.col-sm-12.no-print > button.br-button.small.primary.side-button")
                navegador.execute_script("arguments[0].click();", botaoPesquisarVeiculo)

                time.sleep(3)

                try:

                    campoCarroPesquisado = WebDriverWait(navegador, 10).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-infracao > app-infracoes-list > app-infracoes-veiculo-list > div > div > app-infracao-veiculo-lista > form > div.ng-star-inserted > div.ng-star-inserted > div:nth-child(2) > div"))
                    ).click()     

                    #Verifica erro ao pesquisar as infrações do carro
                    try:
                        erro = WebDriverWait(navegador, 10).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "br-main-layout main > br-alert-messages div.content > div"))
                        )
                        mensagem_erro = erro.text

                        if "Ocorreu erro na validação do campo abaixo:" == mensagem_erro:
                            time.sleep(10)
                            navegador.refresh()

                    except (TimeoutException, NoSuchElementException):
                        pass  # Se não houver erro, continua normalmente

                    campoInfracoesPorVeiculo = WebDriverWait(navegador, 15).until(
                        EC.visibility_of_element_located((By.XPATH, "/html/body/app-root/form/br-main-layout/div/div/main/app-infracao/app-infracoes-list/app-infracoes-veiculo-list/app-infrator-list/div/div/br-title/div/div[1]/h1"))
                    )

                    #--------------------------------- TRATIVAS MULTAS A VENCER --------------------------------#
                    
                    elementoInfraVencer = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-infracao > app-infracoes-list > app-infracoes-veiculo-list > app-infrator-list > div > div > app-infracao-lista > form > div.lista-infracoes.ng-star-inserted > div.ng-star-inserted > h6")
                    textoInfraAvencer = elementoInfraVencer.text
                    regex_palavras_indesejadas = r'\((\d+)\)'
                    resultado_infra_aVencer = re.findall(regex_palavras_indesejadas, textoInfraAvencer)
                    quant_Infra_aVencer = int (resultado_infra_aVencer[0])

                    if quant_Infra_aVencer > 1:
                        
                        for cont in range(1, quant_Infra_aVencer):

                            campo_exibir = navegador.find_element(By.CSS_SELECTOR, "div.ng-star-inserted > div:nth-child(7) > div > br-pagination-table > div > br-select:nth-child(2) div.ng-input > input[type=text]")
                            campo_exibir.send_keys("50")
                            time.sleep(1.5)
                            campo_exibir.send_keys(Keys.ENTER)

                            selector_infracoes_aVencer = f"/html/body/app-root/form/br-main-layout/div/div/main/app-infracao/app-infracoes-list/app-infracoes-veiculo-list/app-infrator-list/div/div/app-infracao-lista/form/div[3]/div[2]/div[{cont}]/div"
                            campo_infracoes_aVencer = navegador.find_element(By.XPATH, selector_infracoes_aVencer)
                            navegador.execute_script("arguments[0].click();", campo_infracoes_aVencer)

                            try:

                                modal_descricao_infra_aVencer = WebDriverWait(navegador, 20).until(
                                    EC.visibility_of_element_located((By.CSS_SELECTOR, "app-infracoes-detail > div > div > div:nth-child(3)"))
                                )

                                campo_orgao_autuador = navegador.find_element(By.CSS_SELECTOR, "div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(1) > td")
                                valor_orgao_autuador = campo_orgao_autuador.text

                                campo_local_infracao = navegador.find_element(By.CSS_SELECTOR, "app-infracoes-detail > div > div > div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(3) > td")
                                valor_local_infracao = campo_local_infracao.text

                                campo_data_hora_cometimentoInfracao = navegador.find_element(By.CSS_SELECTOR, "app-infrator-list > app-infracoes-detail > div > div > div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(4) > td")
                                valor_data_hora_cometimentoInfracao = campo_data_hora_cometimentoInfracao.text

                                campo_AIT = navegador.find_element(By.CSS_SELECTOR, "app-infrator-list > app-infracoes-detail > div > div > div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(5) > td")
                                valor_AIT = campo_AIT.text

                                campo_codigo_infracao = navegador.find_element(By.CSS_SELECTOR, "app-infracoes-list > app-infracoes-veiculo-list > app-infrator-list > app-infracoes-detail > div > div > div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(6) > td")
                                valor_codigo_infracao = campo_codigo_infracao.text

                                campo_valor_infracao = navegador.find_element(By.CSS_SELECTOR, "app-infrator-list > app-infracoes-detail > div > div > div:nth-child(3) > div > table.table.termos > tbody > tr:nth-child(8) > td")
                                valor_infracao = campo_valor_infracao.text

                                colunaAuto = set(row.value for row in guia_resultado_autos ['A'] if row.value is not None)
                                valor_AIT_procurar = valor_AIT

                                existe = valor_AIT_procurar in colunaAuto

                                if not existe:
                                    
                                    ultima_linha =  guia_resultado_autos.max_row + 1

                                    guia_resultado_autos[f'A{ultima_linha}'] = valor_AIT
                                    guia_resultado_autos[f'{ultima_linha}'] = placa_atual


                                    
                                    corte_data_hora_infracao = re.findall(" ", valor_data_hora_cometimentoInfracao)


                            except (TimeoutException, NoSuchElementException):    
                                navegador.refresh()

                            voltarInfra()

                except (TimeoutException, NoSuchElementException):

                    #botaoPesquisarVeiculo = navegador.find_element(By.CSS_SELECTOR, "body > app-root > form > br-main-layout > div > div > main > app-infracao > app-infracoes-list > app-infracoes-veiculo-list > div > div > app-infracao-veiculo-lista > form > div:nth-child(2) > div.col-lg-6.col-md-6.col-sm-12.no-print > button.br-button.small.primary.side-button")
                    #botaoPesquisarVeiculo.click()  
                     
                    pass

                time.sleep(3)

                # Verificar erros durante a pesquisa
                try:
                    erro = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "br-main-layout main > br-alert-messages div.content > div"))
                    )
                    mensagem_erro = erro.text

                    if "Ocorreu erro na validação do campo abaixo:" == mensagem_erro:
                        navegador.refresh()
                        index -= 1
                        continue

                except (TimeoutException, NoSuchElementException):
                    pass  # Se não houver erro, continua normalmente

                # Verificar mensagem de "Não foram encontradas infrações."
                try:
                    campoNaoForamEncontradaInfra = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "br-main-layout main > br-alert-messages div.content > div"))
                    )
                    if "Não foram encontradas infrações." == campoNaoForamEncontradaInfra.text:
                        guia_veiculos[f'C{index + 2}'] = "Não foram encontradas infrações"
                        planilha.save(entradaExcel)
                        voltarPesquisa()
                        continue

                except (TimeoutException, NoSuchElementException):
                    pass

                # Verificar mensagem de "Nenhum registro encontrado"
                try:
                    campoNenhumReg = WebDriverWait(navegador, 10).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, "br-main-layout main > br-alert-messages div.content > div"))
                    )
                    if "Nenhum registro encontrado" == campoNenhumReg.text:
                        guia_veiculos[f'C{index + 2}'] = "Nenhum registro encontrado"
                        planilha.save(entradaExcel)
                        voltarPesquisa()
                        continue

                except (TimeoutException, NoSuchElementException):
                    pass

                time.sleep(2)

            except Exception as e:
                print(f"Erro ao processar a placa {placa_atual}: {e}")

        # Incrementa o índice para a próxima linha
        index += 1

        # Salva a planilha após o processamento completo
        planilha.save(entradaExcel)
        print("Processamento concluído e planilha salva com sucesso.")

        

except TimeoutException:
    print("LOGIN NO SITE NN REALIZADO")

navegador.quit()